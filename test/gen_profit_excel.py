#!/usr/bin/env python3
"""Generate profit comparison Excel workbook."""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
import os

OUT = r'c:\Users\gurs2\OneDrive\Documents\GitHub\telecom-automation\test\Profit_Comparison.xlsx'

# ─── Style helpers ────────────────────────────────────────────────────────────

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=11, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")

def align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def thick_border():
    t = Side(style="medium")
    th = Side(style="thin")
    return Border(left=t, right=t, top=t, bottom=t)

NAVY   = "0A2955"
BLUE   = "175CA6"
TEAL   = "007E8A"
GREEN  = "1B6F42"
ORANGE = "C45008"
PURPLE = "5A238C"
RED    = "B51A1A"
LGRAY  = "F0F2F5"
DGRAY  = "444444"
GOLD   = "F5C518"
WHITE  = "FFFFFF"
LGREEN = "E8F5E9"
LRED   = "FFEBEE"
LYELLOW= "FFFDE7"


def style_header_row(ws, row, cols, bg, fg="FFFFFF", sz=12, bold=True):
    for col in range(1, cols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = fill(bg)
        c.font = font(bold=bold, color=fg, size=sz)
        c.alignment = align()
        c.border = border()


def style_row(ws, row, cols, bg=LGRAY, fg=DGRAY, bold=False, sz=11, alt=False):
    bg = "FFFFFF" if not alt else LGRAY
    for col in range(1, cols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = fill(bg)
        c.font = font(bold=bold, color=fg, size=sz)
        c.alignment = align()
        c.border = border()


def currency(val):
    return f"₹{val:,.0f}"


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 1 — Summary
# ═══════════════════════════════════════════════════════════════════════════════

def sheet_summary(wb):
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [3, 32, 20, 20, 22])

    # Title block
    ws.row_dimensions[1].height = 10
    ws.row_dimensions[2].height = 38
    ws.merge_cells("B2:E2")
    c = ws["B2"]
    c.value = "Telecom AI Automation — Profit Comparison"
    c.fill = fill(NAVY)
    c.font = font(bold=True, color=WHITE, size=18)
    c.alignment = align(h="center")

    ws.row_dimensions[3].height = 22
    ws.merge_cells("B3:E3")
    c = ws["B3"]
    c.value = "Malleswaram 4G/5G NSA Network  ·  30 Cells  ·  3 DUs  ·  18,400 Peak UEs"
    c.fill = fill(BLUE)
    c.font = font(bold=False, color=WHITE, size=11, italic=True)
    c.alignment = align(h="center")

    # Section: Assumptions
    ws.row_dimensions[5].height = 22
    ws.merge_cells("B5:E5")
    c = ws["B5"]
    c.value = "① INPUT ASSUMPTIONS"
    c.fill = fill(TEAL)
    c.font = font(bold=True, color=WHITE, size=12)
    c.alignment = align(h="left")

    headers = ["Metric", "Without AI Model", "With AI Model", "Difference"]
    for i, h in enumerate(headers, 2):
        c = ws.cell(row=6, column=i)
        c.value = h
        c.fill = fill(DGRAY)
        c.font = font(bold=True, color=WHITE, size=11)
        c.alignment = align()
        c.border = border()

    assumptions = [
        ("CAPEX (one-time)",       "₹0",          "₹30,00,000",   "−₹30,00,000"),
        ("Monthly OPEX",           "₹1,50,000",   "₹50,000",      "Save ₹1,00,000/month"),
        ("Users Served",           "1,500",        "2,000",         "+500 users"),
        ("ARPU (per user/month)",  "₹180",        "₹200",          "+₹20/user"),
        ("Monthly Revenue",        "₹2,70,000",   "₹4,00,000",    "+₹1,30,000/month"),
        ("Monthly Profit",         "₹1,20,000",   "₹3,50,000",    "+₹2,30,000/month"),
    ]
    for r, (metric, wo, w, diff) in enumerate(assumptions, 7):
        alt = (r % 2 == 0)
        bg = LGRAY if alt else WHITE
        data = [metric, wo, w, diff]
        for col, val in enumerate(data, 2):
            c = ws.cell(row=r, column=col)
            c.value = val
            c.fill = fill(bg)
            c.font = font(bold=(col == 2), color=DGRAY, size=11)
            c.alignment = align()
            c.border = border()
            if col == 4 and "Save" in str(val) or "+₹" in str(val) or "+500" in str(val):
                c.font = font(bold=True, color=GREEN, size=11)
            if col == 4 and "−" in str(val):
                c.font = font(bold=True, color=RED, size=11)

    # Section: Annual Profit
    r = 14
    ws.row_dimensions[r].height = 22
    ws.merge_cells(f"B{r}:E{r}")
    c = ws[f"B{r}"]
    c.value = "② ANNUAL PROFIT COMPARISON"
    c.fill = fill(TEAL)
    c.font = font(bold=True, color=WHITE, size=12)
    c.alignment = align(h="left")

    for i, h in enumerate(["Metric", "Without AI Model", "With AI Model", "Difference"], 2):
        c = ws.cell(row=r+1, column=i)
        c.value = h
        c.fill = fill(DGRAY)
        c.font = font(bold=True, color=WHITE, size=11)
        c.alignment = align()
        c.border = border()

    annual = [
        ("Annual Revenue",      "₹32,40,000",  "₹48,00,000",  "+₹15,60,000"),
        ("Annual OPEX",         "₹18,00,000",  "₹6,00,000",   "Save ₹12,00,000"),
        ("Annual Gross Profit", "₹14,40,000",  "₹42,00,000",  "+₹27,60,000"),
    ]
    for i, (metric, wo, w, diff) in enumerate(annual, r+2):
        alt = (i % 2 == 0)
        bg = LGRAY if alt else WHITE
        bold_row = (metric == "Annual Gross Profit")
        for col, val in enumerate([metric, wo, w, diff], 2):
            c = ws.cell(row=i, column=col)
            c.value = val
            c.fill = fill(LGREEN if bold_row else bg)
            c.font = font(bold=bold_row, color=DGRAY if not bold_row else GREEN, size=11)
            c.alignment = align()
            c.border = border()
            if col == 4 and bold_row:
                c.font = font(bold=True, color=GREEN, size=12)

    # Section: Key Takeaways
    r = 19
    ws.row_dimensions[r].height = 22
    ws.merge_cells(f"B{r}:E{r}")
    c = ws[f"B{r}"]
    c.value = "③ KEY TAKEAWAYS"
    c.fill = fill(TEAL)
    c.font = font(bold=True, color=WHITE, size=12)
    c.alignment = align(h="left")

    for i, h in enumerate(["Metric", "Value", "", ""], 2):
        c = ws.cell(row=r+1, column=i)
        c.value = h
        c.fill = fill(DGRAY)
        c.font = font(bold=True, color=WHITE, size=11)
        c.alignment = align()
        c.border = border()

    takeaways = [
        ("CAPEX Payback Period",             "~9 months"),
        ("Annual Profit Uplift (from Yr 2)", "₹27,60,000  (+191%)"),
        ("OPEX Savings per Year",            "₹12,00,000"),
        ("Revenue Uplift per Year",          "₹15,60,000"),
        ("5-Year Total Profit (With AI)",    "₹1,80,00,000"),
        ("5-Year Total Profit (Without AI)", "₹72,00,000"),
        ("5-Year Net Gain",                  "₹1,08,00,000"),
    ]
    for i, (metric, val) in enumerate(takeaways, r+2):
        alt = (i % 2 == 0)
        bg = LGRAY if alt else WHITE
        highlight = "5-Year Net Gain" in metric
        c = ws.cell(row=i, column=2)
        c.value = metric
        c.fill = fill(LYELLOW if highlight else bg)
        c.font = font(bold=highlight, color=DGRAY, size=11)
        c.alignment = align()
        c.border = border()
        c = ws.cell(row=i, column=3)
        c.value = val
        c.fill = fill(LYELLOW if highlight else bg)
        c.font = font(bold=highlight, color=GREEN if highlight else DGRAY, size=11 if not highlight else 12)
        c.alignment = align()
        c.border = border()
        for col in [4, 5]:
            c = ws.cell(row=i, column=col)
            c.fill = fill(LYELLOW if highlight else bg)
            c.border = border()

    # Footer
    r = 29
    ws.merge_cells(f"B{r}:E{r}")
    c = ws[f"B{r}"]
    c.value = "AI Model: Gemini 2.5 Flash · KPI Agent (LSTM) · 30 cells · DU-MLS-1/2/3 · Malleswaram, Bangalore"
    c.fill = fill(NAVY)
    c.font = font(bold=False, color="AACCEE", size=9, italic=True)
    c.alignment = align(h="center")


# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 2 — Year-by-Year
# ═══════════════════════════════════════════════════════════════════════════════

def sheet_yearly(wb):
    ws = wb.create_sheet("Year-by-Year")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [3, 14, 22, 22, 22, 22])

    ws.row_dimensions[1].height = 10
    ws.row_dimensions[2].height = 35
    ws.merge_cells("B2:F2")
    c = ws["B2"]
    c.value = "Year-by-Year Net Profit Comparison"
    c.fill = fill(NAVY)
    c.font = font(bold=True, color=WHITE, size=16)
    c.alignment = align()

    # Headers
    headers = ["Year", "Without AI (₹)", "With AI Gross (₹)", "CAPEX Deducted (₹)", "With AI Net (₹)", "Annual Gain (₹)"]
    for i, h in enumerate(headers, 2):
        c = ws.cell(row=4, column=i)
        c.value = h
        c.fill = fill(DGRAY)
        c.font = font(bold=True, color=WHITE, size=11)
        c.alignment = align(wrap=True)
        c.border = border()
    ws.row_dimensions[4].height = 32

    rows = [
        (1, 1440000, 4200000, 3000000, 1200000,  -240000),
        (2, 1440000, 4200000, 0,        4200000, 2760000),
        (3, 1440000, 4200000, 0,        4200000, 2760000),
        (4, 1440000, 4200000, 0,        4200000, 2760000),
        (5, 1440000, 4200000, 0,        4200000, 2760000),
    ]

    wo_cum = 0; w_cum = 0
    for r, (yr, wo, wg, cap, wn, gain) in enumerate(rows, 5):
        alt = (r % 2 == 0)
        bg = LGRAY if alt else WHITE
        wo_cum += wo; w_cum += wn
        data = [f"Year {yr}", wo, wg, cap if cap else "—", wn, gain]
        for col, val in enumerate(data, 2):
            c = ws.cell(row=r, column=col)
            if isinstance(val, int):
                c.value = val
                c.number_format = '₹#,##0'
            else:
                c.value = val
            c.fill = fill(bg)
            c.font = font(bold=(col == 6), color=(GREEN if (isinstance(val, int) and val > 0) else (RED if (isinstance(val, int) and val < 0) else DGRAY)), size=11)
            c.alignment = align()
            c.border = border()

    # Totals row
    r = 10
    ws.row_dimensions[r].height = 24
    totals = ["5-Year TOTAL", 7200000, 21000000, 3000000, 18000000, 10800000]
    for col, val in enumerate(totals, 2):
        c = ws.cell(row=r, column=col)
        if isinstance(val, int):
            c.value = val
            c.number_format = '₹#,##0'
        else:
            c.value = val
        c.fill = fill(NAVY)
        c.font = font(bold=True, color=WHITE, size=12)
        c.alignment = align()
        c.border = border()

    # Note
    ws.merge_cells("B12:F12")
    c = ws["B12"]
    c.value = "Note: Year 1 net gain is negative due to ₹30L CAPEX. From Year 2 onwards AI model delivers ₹27.6L additional profit per year."
    c.font = font(bold=False, color=DGRAY, size=10, italic=True)
    c.alignment = align(h="left")

    # ── Bar Chart ─────────────────────────────────────────────────────────────
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = "Annual Net Profit: With vs Without AI Model"
    chart.y_axis.title = "Profit (₹)"
    chart.x_axis.title = "Year"
    chart.style = 10
    chart.width = 18
    chart.height = 12

    wo_data = Reference(ws, min_col=3, max_col=3, min_row=4, max_row=9)
    ai_data = Reference(ws, min_col=6, max_col=6, min_row=4, max_row=9)
    cats    = Reference(ws, min_col=2, max_col=2, min_row=5, max_row=9)

    chart.add_data(wo_data, titles_from_data=True)
    chart.add_data(ai_data, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.solidFill = "B0BEC5"
    chart.series[0].graphicalProperties.line.solidFill = "78909C"
    chart.series[1].graphicalProperties.solidFill = "1B6F42"
    chart.series[1].graphicalProperties.line.solidFill = "0A2955"

    ws.add_chart(chart, "B14")


# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 3 — Monthly Cashflow
# ═══════════════════════════════════════════════════════════════════════════════

def sheet_monthly(wb):
    ws = wb.create_sheet("Monthly Cashflow")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [3, 12, 18, 18, 18, 18, 18])

    ws.row_dimensions[2].height = 35
    ws.merge_cells("B2:G2")
    c = ws["B2"]
    c.value = "Monthly Cashflow — With AI Model (Year 1 CAPEX Amortised over 12 months)"
    c.fill = fill(NAVY)
    c.font = font(bold=True, color=WHITE, size=14)
    c.alignment = align()

    headers = ["Month", "Revenue (₹)", "OPEX (₹)", "CAPEX (₹)", "Net Profit (₹)", "Cumulative (₹)"]
    for i, h in enumerate(headers, 2):
        c = ws.cell(row=4, column=i)
        c.value = h
        c.fill = fill(DGRAY)
        c.font = font(bold=True, color=WHITE, size=11)
        c.alignment = align(wrap=True)
        c.border = border()
    ws.row_dimensions[4].height = 28

    capex = 3000000
    revenue = 400000
    opex = 50000
    cumulative = -capex

    for m in range(1, 13):
        r = m + 4
        monthly_capex = capex if m == 1 else 0
        net = revenue - opex - monthly_capex
        cumulative += (revenue - opex)
        alt = (m % 2 == 0)
        bg = LGRAY if alt else WHITE
        payback_month = cumulative >= 0
        row_bg = "E8F5E9" if payback_month else bg

        data = [f"Month {m}", revenue, opex, monthly_capex if m == 1 else 0, revenue - opex - monthly_capex, cumulative]
        for col, val in enumerate(data, 2):
            c = ws.cell(row=r, column=col)
            if isinstance(val, (int, float)):
                c.value = val
                c.number_format = '₹#,##0'
            else:
                c.value = val
            c.fill = fill(row_bg)
            fc = GREEN if (isinstance(val, (int, float)) and val > 0 and col > 2) else (RED if (isinstance(val, (int, float)) and val < 0) else DGRAY)
            c.font = font(bold=payback_month and col == 7, color=fc, size=11)
            c.alignment = align()
            c.border = border()

        if payback_month and cumulative >= 0 and cumulative - (revenue - opex) < 0:
            ws.merge_cells(f"H{r}:H{r}")
            note = ws.cell(row=r, column=8)
            note.value = "← PAYBACK ACHIEVED"
            note.fill = fill("1B6F42")
            note.font = font(bold=True, color=WHITE, size=11)
            note.alignment = align(h="left")

    # Totals
    r = 17
    ws.row_dimensions[r].height = 24
    totals_label = ["YEAR 1 TOTAL", 400000*12, 50000*12, 3000000, 400000*12 - 50000*12 - 3000000, 400000*12 - 50000*12 - 3000000]
    for col, val in enumerate(totals_label, 2):
        c = ws.cell(row=r, column=col)
        if isinstance(val, (int, float)):
            c.value = val
            c.number_format = '₹#,##0'
        else:
            c.value = val
        c.fill = fill(NAVY)
        c.font = font(bold=True, color=WHITE, size=12)
        c.alignment = align()
        c.border = border()


# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 4 — ROI Metrics
# ═══════════════════════════════════════════════════════════════════════════════

def sheet_roi(wb):
    ws = wb.create_sheet("ROI Metrics")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [3, 36, 24, 3])

    ws.row_dimensions[2].height = 35
    ws.merge_cells("B2:C2")
    c = ws["B2"]
    c.value = "ROI & Business Case Metrics"
    c.fill = fill(NAVY)
    c.font = font(bold=True, color=WHITE, size=16)
    c.alignment = align()

    sections = [
        ("FINANCIAL RETURNS", TEAL, [
            ("CAPEX",                          "₹30,00,000"),
            ("Annual OPEX (With AI)",          "₹6,00,000"),
            ("Annual OPEX (Without AI)",       "₹18,00,000"),
            ("Annual OPEX Savings",            "₹12,00,000"),
            ("Annual Revenue Uplift",          "₹15,60,000"),
            ("Annual Profit (With AI)",        "₹42,00,000"),
            ("Annual Profit (Without AI)",     "₹14,40,000"),
            ("Annual Profit Uplift",           "₹27,60,000  (+191%)"),
        ]),
        ("PAYBACK & BREAKEVEN", PURPLE, [
            ("CAPEX Payback Period",           "~9 months"),
            ("Cumulative Breakeven vs Manual", "Month 14"),
            ("Monthly cash positive from",     "Month 2 (after CAPEX month)"),
        ]),
        ("5-YEAR OUTLOOK", GREEN, [
            ("5-Year Revenue (With AI)",       "₹2,40,00,000"),
            ("5-Year Revenue (Without AI)",    "₹1,62,00,000"),
            ("5-Year Profit (With AI)",        "₹1,80,00,000"),
            ("5-Year Profit (Without AI)",     "₹72,00,000"),
            ("5-Year Net Gain from AI",        "₹1,08,00,000"),
            ("5-Year ROI",                     "360%"),
        ]),
        ("OPERATIONAL IMPROVEMENTS", ORANGE, [
            ("Additional Users Served",        "+500 users (+33%)"),
            ("ARPU Improvement",               "+₹20/user/month (+11%)"),
            ("OPEX Reduction",                 "67% lower (₹1.5L → ₹0.5L/month)"),
            ("Network Incidents (manual)",     "~3-5 per month (human latency)"),
            ("Network Incidents (AI)",         "<1 per month (auto-resolved in 5s)"),
            ("Avg Resolution Time",            "5 seconds (TOPO_POLL_SEC)"),
        ]),
    ]

    row = 4
    for section_title, color, items in sections:
        ws.row_dimensions[row].height = 24
        ws.merge_cells(f"B{row}:C{row}")
        c = ws[f"B{row}"]
        c.value = section_title
        c.fill = fill(color)
        c.font = font(bold=True, color=WHITE, size=12)
        c.alignment = align(h="left")
        row += 1

        for i, (metric, val) in enumerate(items):
            alt = (i % 2 == 0)
            bg = LGRAY if alt else WHITE
            highlight = any(x in metric for x in ["Net Gain", "ROI", "Profit Uplift", "OPEX Savings"])
            c = ws.cell(row=row, column=2)
            c.value = metric
            c.fill = fill(LYELLOW if highlight else bg)
            c.font = font(bold=highlight, color=DGRAY, size=11)
            c.alignment = align(h="left")
            c.border = border()
            c = ws.cell(row=row, column=3)
            c.value = val
            c.fill = fill(LYELLOW if highlight else bg)
            c.font = font(bold=highlight, color=GREEN if highlight else DGRAY, size=11 if not highlight else 12)
            c.alignment = align()
            c.border = border()
            row += 1
        row += 1

    # Footer
    ws.merge_cells(f"B{row}:C{row}")
    c = ws[f"B{row}"]
    c.value = "Model: Gemini 2.5 Flash + KPI LSTM Agent · Auto load-balance via move_cell · topology.json atomic update · 5s propagation"
    c.fill = fill(NAVY)
    c.font = font(bold=False, color="AACCEE", size=9, italic=True)
    c.alignment = align(h="center")


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    wb = openpyxl.Workbook()
    sheet_summary(wb)
    sheet_yearly(wb)
    sheet_monthly(wb)
    sheet_roi(wb)
    wb.save(OUT)
    print(f"Saved: {OUT}")

if __name__ == '__main__':
    main()
