"""
Validate LSTM KPI classifier + rule-based fallback across:
  NORMAL, OVERLOAD, SINR_LOW (poor signal), POWER_WASTE

Tests 7 synthetic scenarios (NORMAL×2, OVERLOAD×2, SINR_LOW×1, POWER_WASTE×2)
plus a 30-cell topology health snapshot using actual topology.json cell configs.

Output: test/Classifier_Validation.xlsx  (4 sheets)
"""

import sys, os, random, json
import numpy as np
import torch
import torch.nn.functional as F
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

ROOT    = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
KPI_DIR = os.path.join(ROOT, 'agents', 'kpi_agent')
sys.path.insert(0, KPI_DIR)

from model import KPIClassifier, SEQ_LEN, N_FEATURES, LABELS, normalise   # noqa: E402
from train import train_model                                               # noqa: E402

MODEL_PATH = os.path.join(KPI_DIR, 'kpi_model.pt')
TOPO_FILE  = os.path.join(ROOT, 'dev-env', 'config', 'topology.json')
OUT        = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                          'Classifier_Validation.xlsx')

np.random.seed(42); random.seed(42)

# ── rule thresholds (must match kpi_agent.py) ─────────────────────────────────
OL_PRB  = 85.0
UL_PRB  = 20.0
SINR_MIN = 5.0
PW_W    = 500.0
PW_UE   = 15

SINR_BASE = {"n78": 22.0, "n41": 20.0, "n28": 29.0, "B3": 26.0, "B40": 23.0}

# DU load levels for the 30-cell health snapshot (post-rebalancing state)
DU_LOAD = {"DU-MLS-1": 0.88, "DU-MLS-2": 0.72, "DU-MLS-3": 0.62}

# ── styles ────────────────────────────────────────────────────────────────────
CLASS_FILL = {
    "NORMAL":      PatternFill("solid", fgColor="C6EFCE"),
    "OVERLOAD":    PatternFill("solid", fgColor="FFEB9C"),
    "UNDERLOAD":   PatternFill("solid", fgColor="BDD7EE"),
    "SINR_LOW":    PatternFill("solid", fgColor="FFC7CE"),
    "POWER_WASTE": PatternFill("solid", fgColor="E2EFDA"),
}
WHITE  = PatternFill("solid", fgColor="FFFFFF")
HDR_F  = PatternFill("solid", fgColor="0A2955")
SUB_F  = PatternFill("solid", fgColor="175CA6")
ALT_F  = PatternFill("solid", fgColor="F5F8FF")

HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
SUB_FONT = Font(bold=True, color="FFFFFF", size=9)
BOLD     = Font(bold=True, size=10)
STD      = Font(size=10)
OK_FONT  = Font(bold=True, color="00B050", size=10)
ERR_FONT = Font(bold=True, color="C00000", size=10)
CENTRE   = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT     = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def _border():
    s = Side(border_style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def hdr_cell(ws, r, c, val, span=1, fill=HDR_F, font=HDR_FONT):
    cc = ws.cell(row=r, column=c, value=val)
    cc.fill = fill; cc.font = font; cc.alignment = CENTRE; cc.border = _border()
    if span > 1:
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c + span - 1)


def wr(ws, r, c, val, fill=None, font=None, align=CENTRE, fmt=None):
    cc = ws.cell(row=r, column=c, value=val)
    cc.fill = fill or WHITE
    cc.font = font or STD
    cc.alignment = align; cc.border = _border()
    if fmt:
        cc.number_format = fmt


# ── simulator (replicates du_simulator.py tick() formulas) ────────────────────
def sim_tick(cell: dict, load: float, sinr_override=None) -> dict:
    band    = cell.get("band", "n78")
    peak_dl = cell.get("peak_dl_mbps", 3600)
    tx_pw   = cell.get("tx_power_w", 950)
    idle_pw = cell.get("idle_power_w", int(tx_pw * 0.25))
    max_ues = cell.get("max_ues", 700)
    load    = min(max(load, 0.0), 1.0)
    ues     = int(max_ues * load)
    prb     = min(98.0, load * 100 * random.uniform(0.92, 1.08))
    dl_tput = round(prb / 100 * peak_dl * random.uniform(0.82, 1.18), 2)
    sinr    = (SINR_BASE.get(band, 22.0) - load * 15 + random.gauss(0, 2.5)
               if sinr_override is None
               else sinr_override + random.gauss(0, 0.8))
    power   = max(idle_pw * 0.90,
                  round(idle_pw + load * (tx_pw - idle_pw)
                        + random.gauss(0, tx_pw * 0.025), 1))
    pkt     = round(max(0.0, (load - 0.75) * 2.5 + random.gauss(0, 0.05)), 3)
    return dict(prb_dl_pct=round(prb, 1), sinr_db=round(sinr, 1),
                connected_ues=ues, power_w=power,
                packet_loss_pct=pkt, dl_throughput_mbps=dl_tput)


def manual_tick(kpis: dict) -> dict:
    keys = ["prb_dl_pct", "sinr_db", "connected_ues",
            "power_w", "packet_loss_pct", "dl_throughput_mbps"]
    arr = np.array([kpis[k] for k in keys], dtype=float)
    arr = arr + np.random.randn(N_FEATURES) * arr * 0.04
    arr = np.clip(arr, 0, None)
    return dict(zip(keys, arr.tolist()))


def sim_seq(cell, load, sinr_override=None):
    return [sim_tick(cell, load + (i - 2) * 0.01, sinr_override)
            for i in range(SEQ_LEN)]


def manual_seq(kpis: dict):
    return [manual_tick(kpis) for _ in range(SEQ_LEN)]


# ── classification helpers ────────────────────────────────────────────────────
def rule_classify(kpis: dict) -> tuple:
    p, s, u, w = (kpis["prb_dl_pct"], kpis["sinr_db"],
                  kpis["connected_ues"], kpis["power_w"])
    if   p > OL_PRB:               return 1, "OVERLOAD"
    elif p < UL_PRB:               return 2, "UNDERLOAD"
    elif s < SINR_MIN:             return 3, "SINR_LOW"
    elif w > PW_W and u < PW_UE:  return 4, "POWER_WASTE"
    else:                          return 0, "NORMAL"


def lstm_classify(model, seq: list) -> tuple:
    keys = ["prb_dl_pct", "sinr_db", "connected_ues",
            "power_w", "packet_loss_pct", "dl_throughput_mbps"]
    raw = [[row[k] for k in keys] for row in seq]
    x   = torch.tensor([normalise(s) for s in raw],
                       dtype=torch.float32).unsqueeze(0)
    with torch.no_grad():
        probs = F.softmax(model(x), dim=1)[0]
    cls = probs.argmax().item()
    return cls, LABELS[cls], probs[cls].item(), probs.tolist()


# ── run scenario 30 times, majority vote ──────────────────────────────────────
def run_scenario(model, seq_fn, expected_cls: int, n: int = 30) -> dict:
    lstm_v, lstm_c, rule_v, last_seq = [], [], [], None
    for _ in range(n):
        seq = seq_fn()
        last_seq = seq
        cls, _, conf, _ = lstm_classify(model, seq)
        lstm_v.append(cls); lstm_c.append(conf)
        r, _ = rule_classify(seq[-1])
        rule_v.append(r)
    lp = max(set(lstm_v), key=lstm_v.count)
    rp = max(set(rule_v), key=rule_v.count)
    lc = [c for c, v in zip(lstm_c, lstm_v) if v == lp]
    return dict(
        lstm_pred=lp,  lstm_label=LABELS[lp],
        lstm_conf=sum(lc) / len(lc) if lc else 0,
        lstm_vote=lstm_v.count(lp) / n,
        lstm_match=(lp == expected_cls),
        rule_pred=rp,  rule_label=LABELS[rp],
        rule_vote=rule_v.count(rp) / n,
        rule_match=(rp == expected_cls),
        last_seq=last_seq,
    )


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 1 — Scenario Tests
# ─────────────────────────────────────────────────────────────────────────────
def sheet1_scenario_tests(wb, model):
    # Reference cells from topology
    C5G  = {"band": "n78", "peak_dl_mbps": 3600, "tx_power_w": 950,
             "idle_power_w": 237, "max_ues": 900}   # MLS_18C_01
    C4G  = {"band": "B3",  "peak_dl_mbps": 150,  "tx_power_w": 200,
             "idle_power_w": 50,  "max_ues": 250}   # MLS_RWS_03
    CN41 = {"band": "n41", "peak_dl_mbps": 2800, "tx_power_w": 950,
             "idle_power_w": 237, "max_ues": 700}   # MLS_18C_02

    # Ideal POWER_WASTE KPI values (from training distribution, not simulator)
    PW_IDEAL = dict(prb_dl_pct=13.0, sinr_db=22.0, connected_ues=7.0,
                    power_w=840.0, packet_loss_pct=0.01, dl_throughput_mbps=145.0)
    # Sim-realistic POWER_WASTE (5G cell at near-zero load → power stays near idle)
    PW_SIM_CELL = dict(band="n78", peak_dl_mbps=3600, tx_power_w=950,
                       idle_power_w=237, max_ues=900)

    scenarios = [
        dict(name="NORMAL — 5G n78 (load 50%)",
             cell_type="5G n78",
             expected_cls=0, expected_label="NORMAL",
             seq_fn=lambda: sim_seq(C5G, 0.50),
             notes="Moderate PRB, good SINR; typical mid-afternoon operation"),
        dict(name="NORMAL — 4G B3 (load 55%)",
             cell_type="4G B3",
             expected_cls=0, expected_label="NORMAL",
             seq_fn=lambda: sim_seq(C4G, 0.55),
             notes="Legacy macro; SINR higher (B3 base=26 dB), power low"),
        dict(name="OVERLOAD — 5G n78 (load 94%)",
             cell_type="5G n78",
             expected_cls=1, expected_label="OVERLOAD",
             seq_fn=lambda: sim_seq(C5G, 0.94),
             notes="PRB saturated >85%; packet loss rising; needs cell move"),
        dict(name="OVERLOAD — 4G B3 (load 94%)",
             cell_type="4G B3",
             expected_cls=1, expected_label="OVERLOAD",
             seq_fn=lambda: sim_seq(C4G, 0.94),
             notes="4G macro at capacity; lower peak throughput than 5G"),
        dict(name="SINR_LOW — 5G n41 (interference, load 45%)",
             cell_type="5G n41",
             expected_cls=3, expected_label="SINR_LOW",
             seq_fn=lambda: sim_seq(CN41, 0.45, sinr_override=2.0),
             notes="Co-channel interference event; SINR forced to 2 dB "
                   "independent of load; moderate PRB so OVERLOAD rule doesn't fire"),
        dict(name="POWER_WASTE — Ideal (training distribution)",
             cell_type="5G (ideal)",
             expected_cls=4, expected_label="POWER_WASTE",
             seq_fn=lambda: manual_seq(PW_IDEAL),
             notes="mMIMO at 840 W with 7 UEs; matches training data. "
                   "NOT reachable from simulator's linear power model."),
        dict(name="POWER_WASTE — Sim-realistic (load 1%, 5G n78)",
             cell_type="5G n78",
             expected_cls=4, expected_label="POWER_WASTE",
             seq_fn=lambda: sim_seq(PW_SIM_CELL, 0.01),
             notes="Same intent but simulator produces idle_pw ≈ 240 W (<500 W threshold). "
                   "Rule and LSTM both fail — BUG: linear power model cannot produce "
                   "high-power + low-UE combination."),
    ]

    # Run all scenarios
    for s in scenarios:
        s["result"] = run_scenario(model, s["seq_fn"], s["expected_cls"])

    ws = wb.active
    ws.title = "Scenario Tests"
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = "KPI Classifier Validation — NORMAL · OVERLOAD · SINR_LOW (Poor Signal) · POWER_WASTE"
    c.fill = HDR_F; c.font = Font(bold=True, color="FFFFFF", size=13)
    c.alignment = CENTRE; ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:J2")
    c = ws["A2"]
    c.value = ("LSTM bidirectional classifier vs rule-based fallback  ·  "
               "30 runs per scenario  ·  seed=42  ·  MIN_CONFIDENCE=0.70")
    c.fill = SUB_F; c.font = SUB_FONT; c.alignment = CENTRE
    ws.row_dimensions[2].height = 20

    # Column headers
    col_defs = [
        ("A", "Scenario",         38),
        ("B", "Cell Type",        14),
        ("C", "Expected",         14),
        ("D", "LSTM Pred",        14),
        ("E", "LSTM Conf %",      12),
        ("F", "LSTM Vote %",      12),
        ("G", "Rule Pred",        14),
        ("H", "Rule Vote %",      12),
        ("I", "LSTM Pass?",       12),
        ("J", "Notes",            52),
    ]
    for col, title, width in col_defs:
        ws.column_dimensions[col].width = width
        c = ws.cell(row=3, column=ord(col) - 64, value=title)
        c.fill = SUB_F; c.font = SUB_FONT; c.alignment = CENTRE; c.border = _border()
    ws.row_dimensions[3].height = 30

    # Data rows
    for ri, s in enumerate(scenarios, start=4):
        ws.row_dimensions[ri].height = 28
        r = s["result"]
        bg = ALT_F if ri % 2 == 0 else WHITE

        def w(col, val, fill=None, font=None, align=CENTRE):
            wr(ws, ri, ord(col) - 64, val, fill=fill or bg, font=font, align=align)

        w("A", s["name"], align=LEFT)
        w("B", s["cell_type"])
        w("C", s["expected_label"],
          fill=CLASS_FILL.get(s["expected_label"], bg))
        w("D", r["lstm_label"],
          fill=CLASS_FILL.get(r["lstm_label"], bg))
        w("E", round(r["lstm_conf"] * 100, 1))
        w("F", round(r["lstm_vote"] * 100, 0))
        w("G", r["rule_label"],
          fill=CLASS_FILL.get(r["rule_label"], bg))
        w("H", round(r["rule_vote"] * 100, 0))
        w("I", "✓ PASS" if r["lstm_match"] else "✗ FAIL",
          font=OK_FONT if r["lstm_match"] else ERR_FONT)
        w("J", s["notes"], align=LEFT, font=Font(size=9, italic=True))

    # Summary row
    passes = sum(1 for s in scenarios if s["result"]["lstm_match"])
    total  = len(scenarios)
    sr = 4 + total + 1
    ws.row_dimensions[sr].height = 22
    ws.merge_cells(f"A{sr}:H{sr}")
    c = ws.cell(row=sr, column=1,
                value=f"LSTM accuracy: {passes}/{total} scenarios correct "
                      f"({passes/total*100:.0f}%)")
    c.fill = HDR_F; c.font = HDR_FONT; c.alignment = CENTRE; c.border = _border()
    ws.merge_cells(f"I{sr}:J{sr}")
    c = ws.cell(row=sr, column=9,
                value=f"Note: scenario 7 is a known bug — see Validation Issues sheet")
    c.fill = PatternFill("solid", fgColor="FFC7CE")
    c.font = Font(bold=True, size=9, color="C00000")
    c.alignment = CENTRE; c.border = _border()

    # Bar chart — LSTM confidence
    cd = sr + 3
    ws.cell(row=cd, column=1, value="Scenario").font = BOLD
    ws.cell(row=cd, column=2, value="LSTM Conf %").font = BOLD
    for i, s in enumerate(scenarios):
        ws.cell(row=cd + 1 + i, column=1, value=s["name"][:30])
        ws.cell(row=cd + 1 + i, column=2, value=round(s["result"]["lstm_conf"] * 100, 1))

    chart = BarChart()
    chart.type = "col"
    chart.title = "LSTM Confidence % by Scenario"
    chart.style = 10
    chart.y_axis.title = "Confidence %"
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 100
    chart.height = 12; chart.width = 24
    data = Reference(ws, min_col=2, min_row=cd, max_row=cd + total)
    cats = Reference(ws, min_col=1, min_row=cd + 1, max_row=cd + total)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, f"A{cd + total + 2}")

    return scenarios


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 2 — 30-Cell Health Snapshot
# ─────────────────────────────────────────────────────────────────────────────
def sheet2_cell_health(wb, model, topo: dict):
    ws = wb.create_sheet("30-Cell Health")
    ws.sheet_view.showGridLines = False

    # Build cell→DU map
    cell_du = {}
    for du_id, du_cfg in topo["dus"].items():
        for cid in du_cfg["cell_ids"]:
            cell_du[cid] = du_id

    # Title
    ws.merge_cells("A1:P1")
    c = ws["A1"]
    c.value = ("30-Cell Health Snapshot — Post-Rebalancing  ·  "
               "DU-MLS-1 @ 88%  ·  DU-MLS-2 @ 72%  ·  DU-MLS-3 @ 62% load")
    c.fill = HDR_F; c.font = Font(bold=True, color="FFFFFF", size=12)
    c.alignment = CENTRE; ws.row_dimensions[1].height = 34

    ws.merge_cells("A2:P2")
    c = ws["A2"]
    c.value = ("One simulated tick per cell using du_simulator.py formulas  ·  "
               "LSTM classification (6-step sequence) + rule-based fallback")
    c.fill = SUB_F; c.font = SUB_FONT; c.alignment = CENTRE
    ws.row_dimensions[2].height = 18

    col_defs = [
        ("A", "Cell ID",           16),
        ("B", "DU",                12),
        ("C", "Gen",               8),
        ("D", "Band",              8),
        ("E", "Vendor",            12),
        ("F", "PRB DL %",         10),
        ("G", "SINR dB",          10),
        ("H", "UEs",               8),
        ("I", "Power W",          10),
        ("J", "Pkt Loss %",       10),
        ("K", "Tput Mbps",        11),
        ("L", "LSTM Class",       13),
        ("M", "LSTM Conf %",      12),
        ("N", "Rule Class",       13),
        ("O", "LSTM=Rule?",       11),
        ("P", "Action",           28),
    ]
    for col, title, width in col_defs:
        ws.column_dimensions[col].width = width
        c = ws.cell(row=3, column=ord(col) - 64, value=title)
        c.fill = SUB_F; c.font = SUB_FONT; c.alignment = CENTRE; c.border = _border()
    ws.row_dimensions[3].height = 28

    # DU subtotal counters
    du_stats = {du: {"overload": 0, "sinr_low": 0, "normal": 0,
                     "underload": 0, "power_waste": 0}
                for du in topo["dus"]}

    cells_sorted = sorted(topo["cells"].items(),
                          key=lambda x: (cell_du.get(x[0], ""), x[0]))
    row = 4
    current_du = None
    class_count = {lbl: 0 for lbl in LABELS.values()}

    for cell_id, cfg in cells_sorted:
        du_id  = cell_du.get(cell_id, "?")
        load   = DU_LOAD.get(du_id, 0.70)

        # DU separator row
        if du_id != current_du:
            current_du = du_id
            ws.row_dimensions[row].height = 18
            ws.merge_cells(f"A{row}:P{row}")
            c = ws.cell(row=row, column=1,
                        value=f"  {du_id}  ·  {len(topo['dus'][du_id]['cell_ids'])} cells  "
                              f"·  load = {int(load*100)}%")
            c.fill = PatternFill("solid", fgColor="D9E1F2")
            c.font = Font(bold=True, size=10, color="0A2955")
            c.alignment = LEFT; c.border = _border()
            row += 1

        # Generate sequence and classify
        seq = sim_seq(cfg, load)
        kpi = seq[-1]  # snapshot tick
        lstm_cls, lstm_lbl, lstm_conf, _ = lstm_classify(model, seq)
        rule_idx, rule_lbl = rule_classify(kpi)
        agree = lstm_lbl == rule_lbl

        # Recommended action
        if lstm_lbl == "OVERLOAD":
            action = "Move cell to lighter DU (auto-rebalance)"
        elif lstm_lbl == "SINR_LOW":
            action = "CRITICAL alert — check interference / tilt"
        elif lstm_lbl == "UNDERLOAD":
            action = "Sleep candidate — flag for energy saving"
        elif lstm_lbl == "POWER_WASTE":
            action = "WARNING — reduce RF power / enable sleep"
        else:
            action = "—"

        class_count[lstm_lbl] = class_count.get(lstm_lbl, 0) + 1
        du_stats[du_id][lstm_lbl.lower()] = \
            du_stats[du_id].get(lstm_lbl.lower(), 0) + 1

        ws.row_dimensions[row].height = 22
        bg = ALT_F if row % 2 == 0 else WHITE

        def w(col, val, fill=None, font=None, align=CENTRE, fmt=None):
            wr(ws, row, ord(col) - 64, val, fill=fill or bg,
               font=font, align=align, fmt=fmt)

        w("A", cell_id, font=BOLD, align=LEFT)
        w("B", du_id)
        w("C", cfg.get("generation", "?"))
        w("D", cfg.get("band", "?"))
        w("E", cfg.get("vendor", "?"))
        w("F", round(kpi["prb_dl_pct"], 1))
        w("G", round(kpi["sinr_db"], 1))
        w("H", kpi["connected_ues"])
        w("I", round(kpi["power_w"], 0))
        w("J", round(kpi["packet_loss_pct"], 3))
        w("K", round(kpi["dl_throughput_mbps"], 0))
        w("L", lstm_lbl, fill=CLASS_FILL.get(lstm_lbl, bg))
        w("M", round(lstm_conf * 100, 1))
        w("N", rule_lbl, fill=CLASS_FILL.get(rule_lbl, bg))
        w("O", "✓" if agree else "✗",
          font=OK_FONT if agree else ERR_FONT)
        w("P", action, align=LEFT, font=Font(size=9))
        row += 1

    # Class distribution summary
    row += 1
    ws.row_dimensions[row].height = 22
    ws.merge_cells(f"A{row}:E{row}")
    c = ws.cell(row=row, column=1, value="Class Distribution (LSTM)")
    c.fill = HDR_F; c.font = HDR_FONT; c.alignment = CENTRE; c.border = _border()
    for ci, (lbl, cnt) in enumerate(class_count.items(), start=6):
        ws.cell(row=row, column=ci, value=f"{lbl}: {cnt}").font = BOLD
    row += 2

    return du_stats


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 3 — KPI Sequence Detail
# ─────────────────────────────────────────────────────────────────────────────
def sheet3_kpi_detail(wb, scenarios: list):
    ws = wb.create_sheet("KPI Sequence Detail")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "KPI Sequence Detail — 6-timestep input fed to the LSTM (last run, seed=42)"
    c.fill = HDR_F; c.font = Font(bold=True, color="FFFFFF", size=12)
    c.alignment = CENTRE; ws.row_dimensions[1].height = 30

    col_widths = {"A": 8, "B": 14, "C": 10, "D": 10, "E": 10,
                  "F": 11, "G": 12, "H": 15}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    row = 2
    for s in scenarios:
        seq = s["result"]["last_seq"]
        if seq is None:
            continue

        # Scenario header
        ws.row_dimensions[row].height = 24
        ws.merge_cells(f"A{row}:H{row}")
        c = ws.cell(row=row, column=1,
                    value=f">> {s['name']}  |  Expected: {s['expected_label']}  "
                          f"->  LSTM: {s['result']['lstm_label']} "
                          f"({s['result']['lstm_conf']*100:.1f}%)")
        c.fill = SUB_F; c.font = SUB_FONT; c.alignment = LEFT; c.border = _border()
        row += 1

        # Column headers
        for ci, title in enumerate(["Step", "PRB DL %", "SINR dB",
                                     "UEs", "Power W",
                                     "Pkt Loss %", "Tput Mbps", "LSTM Norm PRB"], 1):
            cc = ws.cell(row=row, column=ci, value=title)
            cc.fill = PatternFill("solid", fgColor="D9E1F2")
            cc.font = Font(bold=True, size=9)
            cc.alignment = CENTRE; cc.border = _border()
        ws.row_dimensions[row].height = 20
        row += 1

        keys = ["prb_dl_pct", "sinr_db", "connected_ues",
                "power_w", "packet_loss_pct", "dl_throughput_mbps"]
        for step_i, tick in enumerate(seq):
            raw = [tick[k] for k in keys]
            norm_prb = normalise(raw)[0]  # just PRB normalised as example
            bg = ALT_F if step_i % 2 == 0 else WHITE
            ws.row_dimensions[row].height = 18
            for ci, val in enumerate([step_i + 1,
                                       round(tick["prb_dl_pct"], 1),
                                       round(tick["sinr_db"], 1),
                                       tick["connected_ues"],
                                       round(tick["power_w"], 0),
                                       round(tick["packet_loss_pct"], 3),
                                       round(tick["dl_throughput_mbps"], 0),
                                       round(norm_prb, 3)], 1):
                wr(ws, row, ci, val, fill=bg)
            row += 1
        row += 1  # blank line between scenarios


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 4 — Validation Issues
# ─────────────────────────────────────────────────────────────────────────────
def sheet4_issues(wb):
    ws = wb.create_sheet("Validation Issues")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "KPI Classifier — Validation Issues Found"
    c.fill = HDR_F; c.font = Font(bold=True, color="FFFFFF", size=13)
    c.alignment = CENTRE; ws.row_dimensions[1].height = 34

    col_defs = [
        ("A", "Issue ID",     10),
        ("B", "Severity",     12),
        ("C", "Component",    22),
        ("D", "Description",  52),
        ("E", "Impact",       36),
        ("F", "Recommended Fix", 52),
    ]
    for col, title, width in col_defs:
        ws.column_dimensions[col].width = width
        c = ws.cell(row=2, column=ord(col) - 64, value=title)
        c.fill = SUB_F; c.font = SUB_FONT; c.alignment = CENTRE; c.border = _border()
    ws.row_dimensions[2].height = 28

    issues = [
        ("BUG-01", "HIGH",
         "kpi_agent.py POWER_WASTE rule",
         "POWER_WASTE_W threshold = 500 W. DU simulator uses linear power model: "
         "power_w = idle_pw + load × (tx_pw − idle_pw). A 5G cell at near-zero load "
         "produces power ≈ 240 W (idle_power_w), far below 500 W. The rule can "
         "never fire for any cell in the simulated fleet.",
         "POWER_WASTE rule is dead code in production. No power-waste alerts will "
         "ever be raised by the rule-based path, even for a cell with 1 UE.",
         "Lower threshold to idle_power_w × 1.2 (≈ 285 W for 5G cells), or make "
         "it relative: 'power_w > idle_power_w × 1.15 AND ues < 15'."),

        ("BUG-02", "HIGH",
         "train.py CLASS 4 training distribution",
         "POWER_WASTE training spec: mean power = 850 W, mean UEs = 7. The simulator "
         "cannot generate this combination — at 7 UEs (load ≈ 0.008), power ≈ 242 W. "
         "The LSTM is trained on a region of feature space that is physically "
         "unreachable from the simulator's linear power model.",
         "LSTM POWER_WASTE class is permanently inaccessible in inference from "
         "simulator data. Classifier will mispredict low-load 5G cells as UNDERLOAD "
         "or NORMAL, never POWER_WASTE. Observed in Scenario 7 (✗ FAIL).",
         "Either (a) change training spec Class 4 to match simulator idle behavior "
         "(mean power ≈ 250 W, ues < 10), or (b) update simulator to model mMIMO "
         "always-on RF power (real hardware holds near-peak power when beams are "
         "configured even with no UEs)."),

        ("BUG-03", "MEDIUM",
         "du_simulator.py — no standalone SINR_LOW path",
         "SINR degradation is modelled only as a function of load: "
         "sinr_db = sinr_base − load × 15. There is no interference injection. "
         "A cell can only reach SINR < 5 dB at very high load (e.g., n78 at "
         "load ≈ 1.13, beyond range, or n41 at load ≈ 1.0: 20 − 15 = 5 dB). "
         "The SINR_LOW class in training (prb=52%, sinr=1 dB) represents "
         "co-channel interference at moderate load — a pattern the simulator "
         "never produces independently.",
         "SINR_LOW and OVERLOAD are conflated in simulator output. At high load "
         "both PRB and SINR degrade together, making it ambiguous which class "
         "to predict. In real deployments interference is load-independent. "
         "SINR_LOW alerts from the LSTM may be suppressed by the OVERLOAD vote.",
         "Add an interference injection mode to the DU simulator: "
         "sinr_db = sinr_base − load×15 − interference_db (env-var per cell). "
         "This allows SINR_LOW scenarios at moderate load, matching training data."),

        ("WARN-04", "LOW",
         "kpi_agent.py Flux query pivot after move_cell",
         "query_latest_cell_kpis() pivots on rowKey ['cell_id','area','du_id','cu_id']. "
         "After move_cell, the 3-minute InfluxDB window contains cell_kpi records "
         "with two different du_id values for the same cell (before and after move). "
         "pivot() treats them as separate rows; the agent may analyse the same "
         "physical cell twice per cycle with different DU contexts.",
         "Minor: duplicate classification of moved cells in the cycle immediately "
         "after a move. Both rows trigger actions. The second action (stale du_id) "
         "may write a misleading alert to InfluxDB.",
         "Add '|> last()' before pivot, keyed on cell_id only, to ensure one row "
         "per cell. Or use range(start: -15s) to reduce the window to one interval."),
    ]

    sev_fill = {
        "HIGH":   PatternFill("solid", fgColor="FFC7CE"),
        "MEDIUM": PatternFill("solid", fgColor="FFEB9C"),
        "LOW":    PatternFill("solid", fgColor="C6EFCE"),
    }
    for ri, (iid, sev, comp, desc, impact, fix) in enumerate(issues, start=3):
        ws.row_dimensions[ri].height = 90
        bg = ALT_F if ri % 2 == 0 else WHITE
        for ci, val in enumerate([iid, sev, comp, desc, impact, fix], 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill = sev_fill[sev] if ci == 2 else bg
            c.font = Font(bold=(ci <= 3), size=9 if ci >= 4 else 10)
            c.alignment = Alignment(horizontal="left" if ci >= 3 else "center",
                                    vertical="top", wrap_text=True)
            c.border = _border()

    # Legend
    row = len(issues) + 4
    ws.merge_cells(f"A{row}:F{row}")
    c = ws.cell(row=row, column=1, value="Severity: HIGH = data quality or silent failure  ·  "
                                          "MEDIUM = partial detection gap  ·  LOW = edge case / minor")
    c.fill = PatternFill("solid", fgColor="F2F2F2")
    c.font = Font(italic=True, size=9)
    c.alignment = CENTRE; c.border = _border()


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────
def main():
    import logging
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")

    print("Loading model …")
    model = KPIClassifier()
    if os.path.exists(MODEL_PATH):
        model.load_state_dict(
            torch.load(MODEL_PATH, map_location="cpu", weights_only=True))
        print(f"  Loaded weights from {MODEL_PATH}")
    else:
        print("  No weights found — training from scratch …")
        model = train_model(MODEL_PATH)
    model.eval()

    print("Loading topology …")
    with open(TOPO_FILE) as f:
        topo = json.load(f)

    print("Building workbook …")
    wb = openpyxl.Workbook()

    print("  Sheet 1: Scenario Tests …")
    scenarios = sheet1_scenario_tests(wb, model)

    print("  Sheet 2: 30-Cell Health Snapshot …")
    sheet2_cell_health(wb, model, topo)

    print("  Sheet 3: KPI Sequence Detail …")
    sheet3_kpi_detail(wb, scenarios)

    print("  Sheet 4: Validation Issues …")
    sheet4_issues(wb)

    wb.save(OUT)
    print(f"\nSaved -> {OUT}")

    # Print summary to console
    print("\n-- Scenario Test Summary ------------------------------")
    for s in scenarios:
        r = s["result"]
        tick = "PASS" if r["lstm_match"] else "FAIL"
        print(f"  [{tick}] {s['name']:<46} "
              f"LSTM={r['lstm_label']:<12} conf={r['lstm_conf']*100:5.1f}%  "
              f"rule={r['rule_label']}")


if __name__ == "__main__":
    main()
